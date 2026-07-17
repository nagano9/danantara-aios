#!/usr/bin/env python3
"""Ingest the Danantara AIOS Actual Condition Input Workbook as a Tier 2 source.

    python 06-scripts/ingest_actual_condition.py <path-to-workbook.xlsx>

WHY THIS EXISTS
---------------
SOURCE_REGISTER.md Tier 2 (Danantara charters, delegations, investment policy) is
empty, so `investment-mandate-interpreter` correctly answers "indeterminate" to
almost every real question. The workbook is that missing tier: 05_DECISION_RIGHTS_DOA
is the delegation matrix, 06_COMMITTEE_CHARTERS the charters, 07_POLICY_REGULATION
the policy register, 15_SKILL_RACI the named owners.

A filled workbook that no skill can read is an archive. This makes it a source.

CLASSIFICATION — READ BEFORE RUNNING
------------------------------------
The workbook's own 99_LOOKUPS offers Confidential, Market-Sensitive, and
Sovereign-Sensitive. A filled workbook may carry all three. It therefore MUST NOT
be committed to this repository, which has no classification boundary of its own.

This script reads the workbook from a path OUTSIDE the repo and, by default, emits
only STRUCTURE and COMPLETENESS: which IDs exist, whether each carries evidence,
who owns it, what is missing. It does not copy threshold values, names, or
document contents into the repo. `--emit-values` exists for use inside an approved
environment; the default is the safe one.

WHAT IT ENFORCES
----------------
The workbook's own rule (00_PETUNJUK, ATURAN EVIDENCE): every material answer must
have a source. This script applies it literally. A DoA row with an approver and a
threshold but no `Clause / Page` and no `Effective Date` is not a delegation — it
is a claim about a delegation, and a skill relying on it would be inventing
authority with extra steps. Those rows are reported as UNSOURCED and are not
promoted to Tier 2.

Requires openpyxl. Deliberately NOT part of the CI gate set: CI is stdlib-only and
there is no workbook in the repo to ingest.
"""
from pathlib import Path
import argparse
import json
import sys

try:
    import openpyxl
except ImportError:
    print("FAILED: openpyxl is required.  pip install openpyxl", file=sys.stderr)
    sys.exit(2)

ROOT = Path(__file__).resolve().parents[1]

HEADER_ROW = 7  # 00_PETUNJUK layout: R1-R3 titles/legend, R7 headers, R8+ data

# Sheets promoted to Tier 2, with the columns that make a row a SOURCE rather
# than a claim. Absent any of these, the row is unsourced.
TIER2_SHEETS = {
    "02_ENTITAS_MANDAT": {
        "id": "ID",
        "label": "Entity / Function",
        "evidence": ["Legal Basis", "Evidence / Source"],
        "owner": "Primary Accountable Owner",
    },
    "05_DECISION_RIGHTS_DOA": {
        "id": "ID",
        "label": "Decision Category",
        # A delegation without clause + effective date is not a delegation.
        "evidence": ["DoA / Legal Document", "Clause / Page", "Effective Date"],
        "owner": "Approver",
    },
    "06_COMMITTEE_CHARTERS": {
        "id": "ID",
        "label": "Committee / Forum Category",
        "evidence": ["Charter / Legal Source", "Effective Date"],
        "owner": "Chair",
    },
    "07_POLICY_REGULATION": {
        "id": "ID",
        "label": "Policy / Standard Category",
        # A policy is citable only if it can be retrieved and dated. A category
        # name with no document, date, or location is a placeholder for a policy,
        # not a policy.
        "evidence": ["Actual Document Name", "Effective Date", "Source URL / Repository"],
        "owner": "Owner",
    },
}

EMPTY = {"", "0", "none", "tbd", "n/a", "-"}

# Open change requests against the workbook, checked on every ingest so they
# cannot quietly lapse. See 08-sources/WORKBOOK_CHANGE_REQUEST.md.
CHANGE_REQUESTS = [
    {
        "id": "CR-001",
        "sheet": "20_DECISION_LOG",
        "columns": [
            "Predicted Outcome / Falsification Trigger",
            "Prediction Verdict",
        ],
        "blocks": "decision-quality-gate, post-decision-learning",
        "why": (
            "decision-quality-gate requires a recorded prediction before PASS, and "
            "post-decision-learning can only settle one recorded BEFORE the "
            "decision. Without these columns the gate correctly RETURNs every "
            "material decision, and every decision is permanently `unlearnable`. "
            "Key Assumptions is a belief, not a scoreable claim."
        ),
    },
]


def is_blank(v) -> bool:
    return v is None or str(v).strip().lower() in EMPTY


def read_sheet(ws):
    """Return (headers, rows) using the workbook's R7-header convention."""
    rows = list(ws.iter_rows(min_row=HEADER_ROW, values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return headers, rows[1:]


def col(headers, name):
    return headers.index(name) if name in headers else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", type=Path)
    ap.add_argument(
        "--emit-values",
        action="store_true",
        help="Include cell values in the output. Approved environments only: the "
        "workbook may carry Confidential / Market-Sensitive / Sovereign-Sensitive "
        "content and this repository has no classification boundary.",
    )
    ap.add_argument("--out", type=Path, default=ROOT / "08-sources" / "tier2_index.json")
    args = ap.parse_args()

    if not args.workbook.is_file():
        print(f"FAILED: not found: {args.workbook}", file=sys.stderr)
        sys.exit(1)

    try:
        args.workbook.resolve().relative_to(ROOT)
        print(
            "FAILED: the workbook sits inside the repository. It may carry "
            "Sovereign-Sensitive content and this repo has no classification "
            "boundary. Move it outside and re-run.",
            file=sys.stderr,
        )
        sys.exit(1)
    except ValueError:
        pass  # outside the repo, which is what we want

    import warnings

    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(args.workbook, data_only=True)

    index, summary = {}, []

    for sheet, spec in TIER2_SHEETS.items():
        if sheet not in wb.sheetnames:
            summary.append((sheet, 0, 0, 0, "SHEET MISSING"))
            continue

        headers, rows = read_sheet(wb[sheet])
        i_id = col(headers, spec["id"])
        i_owner = col(headers, spec["owner"]) if spec["owner"] else None
        i_ev = [c for c in (col(headers, e) for e in spec["evidence"]) if c is not None]

        entries, sourced, unsourced, empty = [], 0, 0, 0
        for r in rows:
            if i_id is None or i_id >= len(r) or is_blank(r[i_id]):
                continue
            rid = str(r[i_id]).strip()

            # The workbook pre-fills grey reference columns (ID, category label).
            # Counting those as Danantara input would report an untouched
            # template as "answered but unsourced", which overstates progress.
            # A row is answered only if something OUTSIDE the reference columns
            # carries content.
            i_label = col(headers, spec["label"]) if spec["label"] else None
            reference_cols = {i_id, i_label}
            has_any = any(
                not is_blank(r[c])
                for c in range(len(r))
                if c not in reference_cols
            )
            ev_present = [
                spec["evidence"][k]
                for k, c in enumerate(i_ev)
                if c < len(r) and not is_blank(r[c])
            ]
            ev_missing = [e for e in spec["evidence"] if e not in ev_present]

            if not has_any:
                state, empty = "EMPTY", empty + 1
            elif ev_missing:
                state, unsourced = "UNSOURCED", unsourced + 1
            else:
                state, sourced = "SOURCED", sourced + 1

            entry = {
                "id": rid,
                "state": state,
                "missing_evidence": ev_missing,
                "owner_present": bool(
                    i_owner is not None and i_owner < len(r) and not is_blank(r[i_owner])
                ),
            }
            if args.emit_values:
                entry["values"] = {
                    h: str(v).strip()
                    for h, v in zip(headers, r)
                    if h and not is_blank(v)
                }
            entries.append(entry)

        index[sheet] = entries
        summary.append((sheet, sourced, unsourced, empty, ""))

    # --- report -------------------------------------------------------------
    print(f"Workbook: {args.workbook.name}")
    print(f"Mode    : {'VALUES (approved env only)' if args.emit_values else 'STRUCTURE ONLY (safe default)'}")
    print()
    print(f"{'sheet':28s} {'SOURCED':>8s} {'UNSOURCED':>10s} {'EMPTY':>6s}")
    print("-" * 56)
    total_sourced = 0
    for sheet, s, u, e, note in summary:
        total_sourced += s
        print(f"{sheet:28s} {s:8d} {u:10d} {e:6d}  {note}")

    print()
    print("SOURCED   = has every evidence column; promotable to Tier 2.")
    print("UNSOURCED = answered but missing clause/date/evidence. Per the workbook's")
    print("            own ATURAN EVIDENCE this is a claim, not a source. A skill")
    print("            relying on it would invent authority with extra steps.")
    print("EMPTY     = not yet filled. Honest and expected.")

    # --- open change requests -------------------------------------------------
    print()
    print("=" * 56)
    print("OPEN CHANGE REQUESTS (08-sources/WORKBOOK_CHANGE_REQUEST.md)")
    print("=" * 56)
    open_crs = 0
    for cr in CHANGE_REQUESTS:
        sheet = cr["sheet"]
        if sheet not in wb.sheetnames:
            print(f"{cr['id']}  SHEET MISSING: {sheet}")
            open_crs += 1
            continue
        headers, _ = read_sheet(wb[sheet])
        missing = [c for c in cr["columns"] if c not in headers]
        if missing:
            open_crs += 1
            print(f"{cr['id']}  OPEN — {sheet} is missing:")
            for c in missing:
                print(f"           - {c}")
            print(f"           blocks: {cr['blocks']}")
            print(f"           why   : {cr['why']}")
        else:
            print(f"{cr['id']}  CLOSED — {sheet} carries all required columns.")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(index, indent=2, ensure_ascii=False), encoding="utf-8")
    print()
    print(f"Wrote {args.out.relative_to(ROOT)}")

    if total_sourced == 0:
        print()
        print("Tier 2 remains EMPTY: no row carries complete evidence yet.")
        print("investment-mandate-interpreter will correctly answer 'indeterminate'")
        print("to mandate questions until DOA rows carry clause + effective date.")
        print("That is the honest answer, not a defect.")


if __name__ == "__main__":
    main()
